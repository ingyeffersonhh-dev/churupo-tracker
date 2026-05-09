"""
Export Router — Genera reportes mensuales en Excel (.xlsx) o PDF.
Incluye desglose por categoría, totales y estado de presupuestos.
"""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from supabase import Client
from supabase_client import get_supabase
from dependencies import get_current_user_id
from decimal import Decimal
from datetime import datetime
import io

router = APIRouter(prefix="/export", tags=["export"])

# ─── Month names ─────────────────────────────────────────────────────────────
MONTHS = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


def _get_report_data(supabase: Client, user_id: str, month: int, year: int) -> dict:
    """Gather all data needed for the report."""
    month_start = f"{year}-{month:02d}-01"
    if month == 12:
        next_month_start = f"{year + 1}-01-01"
    else:
        next_month_start = f"{year}-{month + 1:02d}-01"

    # Transactions
    txs = (
        supabase.table("transactions")
        .select("*")
        .eq("user_id", user_id)
        .gte("transaction_date", month_start)
        .lt("transaction_date", next_month_start)
        .order("transaction_date", desc=True)
        .execute()
    )

    # Categories
    cats = supabase.table("categories").select("id, name, type").eq("user_id", user_id).execute()
    cat_map = {c["id"]: c for c in cats.data}

    # Budgets
    budgets = (
        supabase.table("budgets")
        .select("*")
        .eq("user_id", user_id)
        .eq("month", month)
        .eq("year", year)
        .execute()
    )
    budget_map = {b["category_id"]: float(b["limit_amount"]) for b in budgets.data}

    # Aggregate by category
    by_category = {}
    total_expenses = Decimal("0")
    total_income = Decimal("0")

    for tx in txs.data:
        usd = tx.get("usd_equivalent")
        if not usd:
            continue
        amount = Decimal(str(usd))
        cat_id = tx.get("category_id") or "uncategorized"
        cat_info = cat_map.get(cat_id, {"name": "Sin categoría", "type": "expense"})

        if cat_info.get("type") == "income":
            total_income += amount
        else:
            total_expenses += amount
            if cat_id not in by_category:
                by_category[cat_id] = {"name": cat_info["name"], "spent": Decimal("0"), "count": 0}
            by_category[cat_id]["spent"] += amount
            by_category[cat_id]["count"] += 1

    categories_list = []
    for cat_id, info in sorted(by_category.items(), key=lambda x: x[1]["spent"], reverse=True):
        budget = budget_map.get(cat_id)
        pct = float(info["spent"] / Decimal(str(budget)) * 100) if budget else None
        categories_list.append({
            "name": info["name"],
            "spent": float(info["spent"]),
            "budget": budget,
            "percentage": round(pct, 1) if pct else None,
            "count": info["count"],
            "status": "Excedido" if pct and pct >= 100 else ("Atención" if pct and pct >= 80 else "OK"),
        })

    return {
        "transactions": txs.data,
        "categories": categories_list,
        "total_expenses": float(total_expenses),
        "total_income": float(total_income),
        "balance": float(total_income - total_expenses),
        "cat_map": cat_map,
        "month_name": MONTHS[month],
        "month": month,
        "year": year,
    }


@router.get("/xlsx")
def export_xlsx(
    month: int | None = None,
    year: int | None = None,
    user_id: str = Depends(get_current_user_id),
    supabase: Client = Depends(get_supabase),
):
    """Export monthly report as Excel (.xlsx) file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    now = datetime.now()
    m = month or now.month
    y = year or now.year
    data = _get_report_data(supabase, user_id, m, y)

    wb = Workbook()

    # ─── Sheet 1: Resumen ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"

    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="1F2937")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    money_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = f"📊 Churupo Tracker — Reporte {data['month_name']} {y}"
    ws["A1"].font = title_font
    ws["A2"] = f"Generado: {now.strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(size=9, italic=True, color="6B7280")

    # KPIs
    ws["A4"] = "Total Gastos:"
    ws["B4"] = data["total_expenses"]
    ws["B4"].number_format = money_fmt
    ws["A4"].font = Font(bold=True)

    ws["A5"] = "Total Ingresos:"
    ws["B5"] = data["total_income"]
    ws["B5"].number_format = money_fmt
    ws["A5"].font = Font(bold=True)

    ws["A6"] = "Balance:"
    ws["B6"] = data["balance"]
    ws["B6"].number_format = money_fmt
    ws["A6"].font = Font(bold=True, color="16A34A" if data["balance"] >= 0 else "DC2626")

    # Category breakdown
    row = 8
    headers = ["Categoría", "Gastado (USD)", "Presupuesto (USD)", "% Usado", "Estado", "# Transacciones"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for cat in data["categories"]:
        row += 1
        ws.cell(row=row, column=1, value=cat["name"]).border = thin_border
        c = ws.cell(row=row, column=2, value=cat["spent"])
        c.number_format = money_fmt
        c.border = thin_border
        c = ws.cell(row=row, column=3, value=cat["budget"] or "—")
        if cat["budget"]:
            c.number_format = money_fmt
        c.border = thin_border
        c = ws.cell(row=row, column=4, value=f"{cat['percentage']}%" if cat["percentage"] else "—")
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        status_cell = ws.cell(row=row, column=5, value=cat["status"])
        status_cell.border = thin_border
        status_cell.alignment = Alignment(horizontal="center")
        if cat["status"] == "Excedido":
            status_cell.font = Font(color="DC2626", bold=True)
        elif cat["status"] == "Atención":
            status_cell.font = Font(color="D97706", bold=True)
        else:
            status_cell.font = Font(color="16A34A")
        ws.cell(row=row, column=6, value=cat["count"]).border = thin_border

    # Auto-width
    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 20

    # ─── Sheet 2: Detalle de Transacciones ───────────────────────────────────
    ws2 = wb.create_sheet("Transacciones")
    headers2 = ["Fecha", "Descripción", "Monto", "Moneda", "Equiv. USD", "Categoría", "Fuente"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, tx in enumerate(data["transactions"], 2):
        cat_id = tx.get("category_id")
        cat_name = data["cat_map"].get(cat_id, {}).get("name", "Sin categoría") if cat_id else "Sin categoría"
        date_str = tx.get("transaction_date", "")[:10]

        ws2.cell(row=i, column=1, value=date_str).border = thin_border
        ws2.cell(row=i, column=2, value=tx.get("description", "")).border = thin_border
        c = ws2.cell(row=i, column=3, value=float(tx.get("amount", 0)))
        c.number_format = money_fmt
        c.border = thin_border
        ws2.cell(row=i, column=4, value=tx.get("currency", "")).border = thin_border
        c = ws2.cell(row=i, column=5, value=float(tx.get("usd_equivalent") or 0))
        c.number_format = money_fmt
        c.border = thin_border
        ws2.cell(row=i, column=6, value=cat_name).border = thin_border
        ws2.cell(row=i, column=7, value=tx.get("source", "")).border = thin_border

    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws2.column_dimensions[col_letter].width = 18

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"churupo_reporte_{data['month_name']}_{y}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/pdf")
def export_pdf(
    month: int | None = None,
    year: int | None = None,
    user_id: str = Depends(get_current_user_id),
    supabase: Client = Depends(get_supabase),
):
    """Export monthly report as PDF."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab not installed")

    now = datetime.now()
    m = month or now.month
    y = year or now.year
    data = _get_report_data(supabase, user_id, m, y)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Title"], fontSize=20, textColor=colors.HexColor("#1F2937"),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle", parent=styles["Normal"], fontSize=10,
        textColor=colors.HexColor("#6B7280"), spaceAfter=20,
    )
    section_style = ParagraphStyle(
        "Section", parent=styles["Heading2"], fontSize=14,
        textColor=colors.HexColor("#6366F1"), spaceBefore=20, spaceAfter=10,
    )

    elements = []

    # Title
    elements.append(Paragraph(f"Churupo Tracker — Reporte {data['month_name']} {y}", title_style))
    elements.append(Paragraph(f"Generado: {now.strftime('%d/%m/%Y %H:%M')}", subtitle_style))

    # KPIs table
    kpi_data = [
        ["Total Gastos", f"${data['total_expenses']:,.2f}"],
        ["Total Ingresos", f"${data['total_income']:,.2f}"],
        ["Balance", f"${data['balance']:,.2f}"],
    ]
    kpi_table = Table(kpi_data, colWidths=[2.5 * inch, 2.5 * inch])
    kpi_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 12),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TEXTCOLOR", (1, 2), (1, 2), colors.HexColor("#16A34A") if data["balance"] >= 0 else colors.HexColor("#DC2626")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
    ]))
    elements.append(kpi_table)

    # Category Breakdown
    elements.append(Paragraph("Desglose por Categoría", section_style))

    if data["categories"]:
        cat_header = ["Categoría", "Gastado", "Presupuesto", "% Usado", "Estado"]
        cat_rows = [cat_header]
        for cat in data["categories"]:
            cat_rows.append([
                cat["name"],
                f"${cat['spent']:,.2f}",
                f"${cat['budget']:,.2f}" if cat["budget"] else "—",
                f"{cat['percentage']}%" if cat["percentage"] else "—",
                cat["status"],
            ])

        cat_table = Table(cat_rows, colWidths=[1.8 * inch, 1.2 * inch, 1.2 * inch, 1 * inch, 1 * inch])
        header_bg = colors.HexColor("#6366F1")
        cat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ]))
        elements.append(cat_table)
    else:
        elements.append(Paragraph("No hay gastos registrados en este período.", styles["Normal"]))

    # Transactions detail
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Detalle de Transacciones", section_style))

    if data["transactions"]:
        tx_header = ["Fecha", "Descripción", "Monto", "Moneda", "USD Equiv."]
        tx_rows = [tx_header]
        for tx in data["transactions"][:50]:  # Limit to 50 to avoid massive PDFs
            tx_rows.append([
                tx.get("transaction_date", "")[:10],
                (tx.get("description", "") or "")[:30],
                f"{float(tx.get('amount', 0)):,.2f}",
                tx.get("currency", ""),
                f"${float(tx.get('usd_equivalent') or 0):,.2f}",
            ])

        tx_table = Table(tx_rows, colWidths=[1.2 * inch, 2.2 * inch, 1.1 * inch, 0.8 * inch, 1.1 * inch])
        tx_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (1, -1), "LEFT"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ]))
        elements.append(tx_table)

        if len(data["transactions"]) > 50:
            elements.append(Spacer(1, 8))
            elements.append(Paragraph(
                f"... y {len(data['transactions']) - 50} transacciones más. Usa el formato Excel para el listado completo.",
                ParagraphStyle("Note", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#9CA3AF")),
            ))

    doc.build(elements)
    buf.seek(0)

    filename = f"churupo_reporte_{data['month_name']}_{y}.pdf"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
